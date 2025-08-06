import streamlit as st
import openpyxl
from openpyxl import load_workbook, Workbook
import io
from copy import copy

st.set_page_config(page_title="Merge Excel Files with Formatting", layout="centered")

st.title("📊 Merge Excel Files with Formatting (1 File, 13 Sheets)")

uploaded_files = st.file_uploader("Upload Excel files", type=["xlsx"], accept_multiple_files=True)

if uploaded_files:
    output = io.BytesIO()
    merged_wb = Workbook()
    # Remove the default empty sheet
    merged_wb.remove(merged_wb.active)

    for file in uploaded_files:
        try:
            # Load uploaded file
            original_wb = load_workbook(file)
            for sheet in original_wb.sheetnames:
                original_ws = original_wb[sheet]
                # Create a new sheet in merged file
                new_sheet_name = file.name[:28]  # Sheet name max = 31, keeping room
                copied_ws = merged_wb.create_sheet(title=new_sheet_name)

                # Copy cell values and formatting
                for row in original_ws.iter_rows():
                    for cell in row:
                        new_cell = copied_ws.cell(row=cell.row, column=cell.col_idx, value=cell.value)
                        if cell.has_style:
                            new_cell.font = copy(cell.font)
                            new_cell.border = copy(cell.border)
                            new_cell.fill = copy(cell.fill)
                            new_cell.number_format = copy(cell.number_format)
                            new_cell.protection = copy(cell.protection)
                            new_cell.alignment = copy(cell.alignment)
        except Exception as e:
            st.error(f"❌ Error processing file {file.name}: {str(e)}")

    # Save final merged workbook
    merged_wb.save(output)
    output.seek(0)

    st.success("✅ Files merged with formatting preserved!")

    st.download_button(
        label="📥 Download Merged Excel File",
        data=output,
        file_name="merged_with_formatting.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
